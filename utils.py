from openpyxl.utils import get_column_letter
from parser import parse_nfe, parse_nfse
from dateutil.parser import parse
from openpyxl import Workbook
from io import BytesIO

def get_text(element, path, ns):
    elem = element.find(path, namespaces=ns)
    return elem.text if elem is not None else None

def parse_date(date):
    try:
        parsed_date = parse(date)
        return parsed_date.strftime("%d/%m/%Y")
    except Exception:
        return date
    
def merge_parsed_data(files, xml_type):
    data = {}
    for file in files:
        if file.filename.lower().endswith(".xml"):
            try:
                if xml_type == "nfe":
                    parsed = parse_nfe(file.read())
                elif xml_type == "nfse":
                    parsed = parse_nfse(file.read())
                else:
                    continue

                for sheet, rows in parsed.items():
                    if isinstance(rows, dict):
                        rows = [rows]
                    if sheet not in data:
                        data[sheet] = []
                    data[sheet].extend(rows)
            except Exception as e:
                raise Exception(f"Erro ao processar o arquivo {file.filename}: {e}")
    return data

def generate_excel(data):
    workbook = Workbook()
    first_sheet = True

    for sheet_name, rows in data.items():
        if isinstance(rows, dict):
            rows = [rows]
        if first_sheet:
            sheet = workbook.active
            sheet.title = sheet_name
            first_sheet = False
        else:
            sheet = workbook.create_sheet(title=sheet_name)
        if not rows:
            continue
        headers = list(rows[0].keys())
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(h, "") for h in headers])
        for i, col in enumerate(sheet.columns, 1):
            max_length = 0
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[get_column_letter(i)].width = max_length + 2

    excel_buffer = BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer