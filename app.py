from flask import Flask, render_template, request, send_file, session
from openpyxl.utils import get_column_letter
from parser import parser_nfe, parser_nfse
from openpyxl import Workbook
from datetime import datetime
from io import BytesIO
import json

app = Flask(__name__)
app.secret_key = "session-key" # Chave secreta para poder utilizar session

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":

        if "xml_files" not in request.files:
            return render_template("error.html", msg="Nenhum arquivo enviado!"), 400

        files = request.files.getlist("xml_files")
        xml_type = request.form.get("xml_type")
        
        data = []
        for file in files:
            if file.filename.lower().endswith(".xml"):
                try:
                    if xml_type == "nfe":
                        data.append(parser_nfe(file.read()))
                    elif xml_type == "nfse":
                        data.extend(parser_nfse(file.read()))
                except Exception as e:
                    return render_template("error.html", msg=f"Erro ao processar o arquivo {file.filename}: {e}"), 500

        if data:
            session["parsed_data"] = json.dumps(data)
            return render_template("results.html", data=data)
        else:
            return render_template("error.html", msg="Nenhum dado válido encontrado!"), 400

    session.pop("parsed_data", None)
    return render_template("index.html")

@app.route("/export", methods=["POST"])
def export():
    if "parsed_data" not in session:
        return render_template("error.html", msg="Nenhum dado para exportar!"), 400

    data = json.loads(session["parsed_data"])

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Geral"

    headers = list(data[0].keys())
    sheet.append(headers)

    for row in data:
        sheet.append(list(row.values()))

    for i, col in enumerate(sheet.columns, 1):
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[get_column_letter(i)].width = max_length + 2

    excel_buffer = BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    
    timestamp = datetime.now().strftime("%d%m%Y_%H%M%S")
    return send_file(
        excel_buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"relatorio_{timestamp}.xlsx"
    )

if __name__ == "__main__":
    app.run(debug=True)